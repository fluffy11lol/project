import os
import time
import openpyxl
import psutil

timee = 0
range = 0.5
while 1:
	timee += range
	available = round(psutil.virtual_memory().available / 1000000)
	used = round(psutil.virtual_memory().used / 1000000)
	usage = psutil.virtual_memory().percent
	path = r".\Result.xlsx"
	file = openpyxl.load_workbook(path)
	sheet = file.active
	sheet.cell(column=1, row=sheet.max_row + 1, value=timee)
	sheet.cell(column=2, row=sheet.max_row, value=used)
	sheet.cell(column=3, row=sheet.max_row, value=usage)
	sheet.cell(column=4, row=sheet.max_row, value=available)
	file.save(path)
	print("_")
	time.sleep(range)
